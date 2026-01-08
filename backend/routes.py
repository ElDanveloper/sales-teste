from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import db
import models
import schemas

router = APIRouter()


# Upload CSV
@router.post("/upload/csv/{file_type}")
async def upload_csv(file_type: str, file: UploadFile = File(...)):
    contents = await file.read()
    try:
        df = pd.read_csv(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido ou corrompido.")
    
    try:
        if file_type == "categories":
            categories_data = []
            for _, row in df.iterrows():
                cat = {
                    "id": int(row['id']),
                    "name": str(row['name'])
                }
                categories_data.append(cat)
            count = db.add_categories_bulk(categories_data)
            return {"message": "Importação de Categorias concluída", "inserted": count}
            
        elif file_type == "products":
            products_data = []
            for _, row in df.iterrows():
                prod = {
                    "id": int(row['id']),
                    "name": str(row['name']),
                    "description": str(row.get('description', '')),
                    "price": float(row['price']),
                    "category_id": int(row['category_id']),
                    "brand": str(row.get('brand', ''))
                }
                products_data.append(prod)
            count = db.add_products_bulk(products_data)
            return {"message": "Importação de Produtos concluída", "inserted": count}

        elif file_type == "sales":
            sales_data = []
            for _, row in df.iterrows():
                try:
                    sale_date = pd.to_datetime(row['date']).strftime('%Y-%m-%d')
                except:
                    continue
                
                sale = {
                    "id": int(row['id']),
                    "product_id": int(row['product_id']),
                    "quantity": int(row['quantity']),
                    "total_price": float(row['total_price']),
                    "date": sale_date
                }
                sales_data.append(sale)
            count = db.add_sales_bulk(sales_data)
            return {"message": "Importação de Vendas concluída", "inserted": count}
            
        else:
            raise HTTPException(status_code=400, detail="Tipo inválido. Use: categories, products, sales")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.get("/products", response_model=list[schemas.ProductResponse])
def list_products():
    products = db.get_products()
    return products


@router.post("/products", response_model=schemas.ProductResponse)
def create_product(product: schemas.ProductCreate):
    products = db.get_products()
    new_id = max([p.get('id', 0) for p in products], default=0) + 1
    product_dict = product.dict()
    product_dict['id'] = new_id
    return db.add_product(product_dict)


@router.delete("/products/{product_id}")
def delete_product(product_id: int):
    db.delete_product(product_id)
    return {"message": "Produto deletado com sucesso"}


@router.get("/categories", response_model=list[schemas.CategoryResponse])
def list_categories():
    categories = db.get_categories()
    return categories


@router.post("/categories", response_model=schemas.CategoryResponse)
def create_category(category: schemas.CategoryCreate):
    categories = db.get_categories()
    new_id = max([c.get('id', 0) for c in categories], default=0) + 1
    category_dict = category.dict()
    category_dict['id'] = new_id
    return db.add_category(category_dict)


@router.get("/sales", response_model=list[schemas.SaleResponse])
def list_sales():
    sales = db.get_sales()
    return sales

@router.post("/sales", response_model=schemas.SaleResponse)
def create_sale(sale: schemas.SaleCreate):
    products = db.get_products()
    if not any(p.get('id') == sale.product_id for p in products):
        raise HTTPException(status_code=400, detail="Produto não encontrado")

    sales = db.get_sales()
    new_id = max([s.get('id', 0) for s in sales], default=0) + 1
    sale_dict = sale.dict()
    sale_dict['id'] = new_id
    return db.add_sale(sale_dict)


@router.put("/sales/{sale_id}", response_model=schemas.SaleResponse)
def update_sale(sale_id: int, sale: schemas.SaleCreate):
    products = db.get_products()
    if not any(p.get('id') == sale.product_id for p in products):
        raise HTTPException(status_code=400, detail="Produto não encontrado")
    
    sale_dict = sale.dict()
    updated = db.update_sale(sale_id, sale_dict)
    if not updated:
        raise HTTPException(status_code=404, detail="Venda não encontrada")
    return updated


@router.get("/dashboard/stats", response_model=schemas.DashboardStats)
def dashboard_stats():
    return db.get_dashboard_stats()


# Exportar relatório Excel
@router.get("/reports/export.xlsx")
def export_xlsx():
    products = db.get_products()
    categories = db.get_categories()
    sales = db.get_sales()
    stats = db.get_dashboard_stats()

    wb = Workbook()
    wb.remove(wb.active)

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Aba Resumo
    ws_summary = wb.create_sheet("Resumo", 0)
    ws_summary['A1'] = "RELATÓRIO SMARTMART SOLUTIONS"
    ws_summary['A1'].font = Font(bold=True, size=14, color="1F4E78")
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'].alignment = center_alignment
    
    ws_summary['A3'] = "Data do Relatório:"
    ws_summary['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws_summary['A5'] = "RESUMO EXECUTIVO"
    ws_summary['A5'].font = Font(bold=True, size=12, color="1F4E78")
    ws_summary.merge_cells('A5:B5')
    
    summary_data = [
        ("Total de Vendas", stats.get('total_sales_count', 0)),
        ("Receita Total", f"R$ {stats.get('total_revenue', 0):,.2f}"),
        ("Produtos Cadastrados", len(products)),
        ("Categorias Ativas", len(categories)),
    ]
    
    row = 6
    for label, value in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        row += 1
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 25
    
    # Aba Produtos
    ws_products = wb.create_sheet("Produtos", 1)
    if products:
        headers_prod = ["ID", "Nome", "Descrição", "Preço", "Marca", "Categoria ID"]
        for col, header in enumerate(headers_prod, 1):
            cell = ws_products.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        for row_idx, prod in enumerate(products, 2):
            ws_products.cell(row=row_idx, column=1).value = prod.get('id')
            ws_products.cell(row=row_idx, column=2).value = prod.get('name')
            ws_products.cell(row=row_idx, column=3).value = prod.get('description', '')
            ws_products.cell(row=row_idx, column=4).value = prod.get('price', 0)
            ws_products.cell(row=row_idx, column=5).value = prod.get('brand', '')
            ws_products.cell(row=row_idx, column=6).value = prod.get('category_id')
            
            for col in range(1, 7):
                cell = ws_products.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = left_alignment
                if col == 4:  # Preço
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = center_alignment
        
        ws_products.column_dimensions['A'].width = 8
        ws_products.column_dimensions['B'].width = 25
        ws_products.column_dimensions['C'].width = 30
        ws_products.column_dimensions['D'].width = 12
        ws_products.column_dimensions['E'].width = 12
        ws_products.column_dimensions['F'].width = 12
    
    # Aba Categorias
    ws_categories = wb.create_sheet("Categorias", 2)
    if categories:
        headers_cat = ["ID", "Nome"]
        for col, header in enumerate(headers_cat, 1):
            cell = ws_categories.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        for row_idx, cat in enumerate(categories, 2):
            ws_categories.cell(row=row_idx, column=1).value = cat.get('id')
            ws_categories.cell(row=row_idx, column=2).value = cat.get('name')
            
            for col in range(1, 3):
                cell = ws_categories.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = left_alignment
        
        ws_categories.column_dimensions['A'].width = 8
        ws_categories.column_dimensions['B'].width = 25
    
    # Aba Vendas
    ws_sales = wb.create_sheet("Vendas", 3)
    if sales:
        headers_sales = ["ID", "Produto ID", "Quantidade", "Total (R$)", "Data"]
        for col, header in enumerate(headers_sales, 1):
            cell = ws_sales.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        for row_idx, sale in enumerate(sales, 2):
            ws_sales.cell(row=row_idx, column=1).value = sale.get('id')
            ws_sales.cell(row=row_idx, column=2).value = sale.get('product_id')
            ws_sales.cell(row=row_idx, column=3).value = sale.get('quantity')
            ws_sales.cell(row=row_idx, column=4).value = sale.get('total_price', 0)
            ws_sales.cell(row=row_idx, column=5).value = sale.get('date')
            
            for col in range(1, 6):
                cell = ws_sales.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = center_alignment
                if col == 4:  # Total
                    cell.number_format = 'R$ #,##0.00'
        
        ws_sales.column_dimensions['A'].width = 8
        ws_sales.column_dimensions['B'].width = 12
        ws_sales.column_dimensions['C'].width = 12
        ws_sales.column_dimensions['D'].width = 15
        ws_sales.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {"Content-Disposition": "attachment; filename=smartmart-report.xlsx"}
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# Exportar produtos CSV
@router.get("/reports/export-products.csv")
def export_products_csv():
    products = db.get_products()
    
    csv_data = "id,name,description,price,category_id,brand\n"
    for prod in products:
        csv_data += f"{prod.get('id')},{prod.get('name')},\"{prod.get('description', '')}\",{prod.get('price')},{prod.get('category_id')},{prod.get('brand', '')}\n"
    
    output = io.BytesIO(csv_data.encode('utf-8'))
    headers = {"Content-Disposition": "attachment; filename=produtos.csv"}
    return StreamingResponse(output, media_type="text/csv", headers=headers)


# Exportar vendas CSV
@router.get("/reports/export-sales.csv")
def export_sales_csv():
    sales = db.get_sales()
    
    csv_data = "id,product_id,quantity,total_price,date\n"
    for sale in sales:
        csv_data += f"{sale.get('id')},{sale.get('product_id')},{sale.get('quantity')},{sale.get('total_price')},{sale.get('date')}\n"
    
    output = io.BytesIO(csv_data.encode('utf-8'))
    headers = {"Content-Disposition": "attachment; filename=vendas.csv"}
    return StreamingResponse(output, media_type="text/csv", headers=headers)


# Postman
@router.get("/postman/collection")
def postman_collection():
    collection = {
        "info": {
            "name": "SmartMart Solutions API",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
            "_postman_id": "smartmart-collection"
        },
        "variable": [{"key": "baseUrl", "value": "http://localhost:3000"}],
        "item": [
            {"name": "List Products", "request": {"method": "GET", "url": "{{baseUrl}}/products"}},
            {"name": "Create Product", "request": {"method": "POST", "header": [{"key": "Content-Type", "value": "application/json"}], "body": {"mode": "raw", "raw": "{\n  \"name\": \"Example\",\n  \"price\": 100.0,\n  \"category_id\": 1\n}"}, "url": "{{baseUrl}}/products"}},
            {"name": "Delete Product", "request": {"method": "DELETE", "url": "{{baseUrl}}/products/1"}},
            {"name": "List Categories", "request": {"method": "GET", "url": "{{baseUrl}}/categories"}},
            {"name": "Create Category", "request": {"method": "POST", "header": [{"key": "Content-Type", "value": "application/json"}], "body": {"mode": "raw", "raw": "{\n  \"name\": \"TVs\"\n}"}, "url": "{{baseUrl}}/categories"}},
            {"name": "List Sales", "request": {"method": "GET", "url": "{{baseUrl}}/sales"}},
            {"name": "Dashboard Stats", "request": {"method": "GET", "url": "{{baseUrl}}/dashboard/stats"}},
            {"name": "Upload CSV - Products", "request": {"method": "POST", "url": "{{baseUrl}}/upload/csv/products"}},
            {"name": "Upload CSV - Categories", "request": {"method": "POST", "url": "{{baseUrl}}/upload/csv/categories"}},
            {"name": "Upload CSV - Sales", "request": {"method": "POST", "url": "{{baseUrl}}/upload/csv/sales"}},
            {"name": "Export XLSX", "request": {"method": "GET", "url": "{{baseUrl}}/reports/export.xlsx"}}
        ]
    }

    headers = {"Content-Disposition": "attachment; filename=SmartMart.postman_collection.json"}
    return JSONResponse(content=collection, headers=headers)